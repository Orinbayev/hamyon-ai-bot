import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.transaction import Transaction

HEADERS = ["ID", "Tur", "Summa", "Valyuta", "Kategoriya", "To'lov", "Izoh", "Sana"]
PAYMENT_MAP = {
    "cash": "Naqd", "card": "Karta", "click": "Click",
    "payme": "Payme", "bank": "Bank", "other": "Boshqa",
}


def _row(t: Transaction) -> list:
    return [
        t.id,
        "Kirim" if t.type == "income" else "Chiqim",
        float(t.amount),
        t.currency,
        t.category.name if t.category else "Boshqa",
        PAYMENT_MAP.get(t.payment_method, t.payment_method),
        t.note or "",
        t.transaction_date.strftime("%d.%m.%Y"),
    ]


def build_excel(transactions: list[Transaction]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tranzaksiyalar"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    inc_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    exp_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[1].height = 22

    for t in transactions:
        ws.append(_row(t))
        fill = inc_fill if t.type == "income" else exp_fill
        for cell in ws[ws.max_row]:
            cell.border = thin
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate([6, 10, 14, 10, 18, 12, 30, 12], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_csv(transactions: list[Transaction]) -> io.BytesIO:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for t in transactions:
        writer.writerow(_row(t))
    result = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    result.seek(0)
    return result
