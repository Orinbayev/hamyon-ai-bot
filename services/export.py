"""
Export service — Excel va CSV eksport.
sync_to_async orqali async bot handler lardan chaqiriladi.
"""

import csv
import io
import logging
from datetime import date
from decimal import Decimal

from asgiref.sync import sync_to_async
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.transactions.models import Transaction
from apps.users.models import TelegramUser

logger = logging.getLogger("services")

HEADERS = ["ID", "Tur", "Summa", "Valyuta", "Kategoriya", "To'lov turi", "Izoh", "Sana", "Yaratilgan"]
PAYMENT_MAP = {"cash": "Naqd", "card": "Karta", "click": "Click", "payme": "Payme", "bank": "Bank", "other": "Boshqa"}


def _get_qs(user: TelegramUser, start: date | None, end: date | None):
    qs = Transaction.objects.filter(user=user).select_related("category").order_by("-transaction_date", "-created_at")
    if start:
        qs = qs.filter(transaction_date__gte=start)
    if end:
        qs = qs.filter(transaction_date__lte=end)
    return list(qs)


def _row(t: Transaction) -> list:
    return [
        t.id,
        "Kirim" if t.type == "income" else "Chiqim",
        float(t.amount),
        t.currency,
        t.category.name if t.category else "Boshqa",
        PAYMENT_MAP.get(t.payment_method, t.payment_method),
        t.note or "",
        t.transaction_date.strftime("%d.%m.%Y"),
        t.created_at.strftime("%d.%m.%Y %H:%M"),
    ]


def _excel_sync(user: TelegramUser, start: date | None = None, end: date | None = None) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tranzaksiyalar"

    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    thin = Border(*[Side(style="thin")] * 0,
                  left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[1].height = 22

    transactions = _get_qs(user, start, end)
    for t in transactions:
        ws.append(_row(t))
        fill = income_fill if t.type == "income" else expense_fill
        for cell in ws[ws.max_row]:
            cell.border = thin
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    for i, w in enumerate([6, 10, 14, 10, 16, 14, 30, 12, 18], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.append([])
    total_income = sum(float(t.amount) for t in transactions if t.type == "income")
    total_expense = sum(float(t.amount) for t in transactions if t.type == "expense")
    for label, val in [("Jami kirim:", total_income), ("Jami chiqim:", total_expense),
                       ("Sof balans:", total_income - total_expense)]:
        ws.append(["", label, val, "UZS", "", "", "", "", ""])
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 3).font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _csv_sync(user: TelegramUser, start: date | None = None, end: date | None = None) -> io.BytesIO:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(HEADERS)
    for t in _get_qs(user, start, end):
        writer.writerow(_row(t))
    result = io.BytesIO(out.getvalue().encode("utf-8-sig"))
    result.seek(0)
    return result


# ── Async public API ─────────────────────────────────────────────────────────

export_excel = sync_to_async(_excel_sync)
export_csv = sync_to_async(_csv_sync)
